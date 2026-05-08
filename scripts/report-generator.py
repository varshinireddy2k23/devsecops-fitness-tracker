import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS = {
    "dependency-report.json": "npm",
    "trivy-report.json": "Trivy",
    "checkov-report.json": "Checkov",
    "gitleaks-report.json": "Gitleaks",
    "opa-report.json": "OPA Policy Gate"
}

output_file = "reports/security-report.xlsx"

summary_data = []

# ---- Read JSON Files ----
for file_name, scanner_name in REPORTS.items():

    path = os.path.join("reports", file_name)

    critical = high = medium = low = 0

    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as file:
                data = json.load(file)

            text = json.dumps(data).lower()

            critical = text.count("critical")
            high = text.count("high")
            medium = text.count("medium")
            low = text.count("low")

        except:
            pass

    total = critical + high + medium + low

    if critical > 0 or high > 5:
        status = "FAIL"
    elif medium > 0 or low > 0:
        status = "WARN"
    else:
        status = "PASS"

    summary_data.append([
        scanner_name,
        critical,
        high,
        medium,
        low,
        status
    ])

# ---- Create Excel ----
df = pd.DataFrame(summary_data, columns=[
    "Scanner", "Critical", "High", "Medium", "Low", "Status"
])

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Summary", index=False)

    ws = writer.book["Summary"]

    # Header Styling
    header_fill = PatternFill("solid", fgColor="000000")
    white_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # Row Colors
    red_fill = PatternFill("solid", fgColor="FF6666")
    orange_fill = PatternFill("solid", fgColor="F4B183")
    green_fill = PatternFill("solid", fgColor="92D050")

    for row in range(2, ws.max_row + 1):
        status = ws[f"F{row}"].value

        if status == "FAIL":
            fill = red_fill
        elif status == "WARN":
            fill = orange_fill
        else:
            fill = green_fill

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill

    # Borders + Auto Width
    thin = Side(style="thin", color="000000")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(
                left=thin, right=thin,
                top=thin, bottom=thin
            )

    for col in ws.columns:
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[
            get_column_letter(col[0].column)
        ].width = length + 5

print("Styled Excel report generated successfully!")